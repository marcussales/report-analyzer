import streamlit as st
import pandas as pd
import pdfplumber
import re
from datetime import datetime, timedelta
import numpy as np
from typing import Dict, List, Tuple, Optional
import io
import warnings
warnings.filterwarnings('ignore')

# CONFIGURAÇÃO DA PÁGINA (DEVE SER O PRIMEIRO COMANDO STREAMLIT)

# Tentar importar openpyxl para geração de Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    st.error("""
    ❌ **Biblioteca openpyxl não encontrada!**
    
    Para gerar relatórios Excel, instale a biblioteca:
    ```
    pip install openpyxl
    ```
    """)
    OPENPYXL_AVAILABLE = False

class PontoComparador:
    """
    Classe principal para comparar relatórios de ponto (PDF) e apropriação (Excel).
    """
    
    def __init__(self):
        self.dados_ponto_pdf = {}
        self.dados_apropriacao_excel = {}
        self.divergencias = []
        self.marcacoes_manuais = {}
    
    def _normalizar_horario(self, horario: str) -> str:
        """
        Normaliza um horário para o formato HH:MM, ignorando segundos.
        """
        if not horario or horario.strip() == '':
            return '00:00'
        
        try:
            horario_limpo = horario.strip()
            
            if re.match(r'^\d{2}:\d{2}$', horario_limpo):
                return horario_limpo
            
            if re.match(r'^\d{2}:\d{2}:\d{2}$', horario_limpo):
                return horario_limpo[:5]
            
            if re.match(r'^\d{1}:\d{2}', horario_limpo):
                return f"0{horario_limpo[:4]}"
            
            for formato in ['%H:%M:%S', '%H:%M', '%H:%M:%S.%f']:
                try:
                    dt = datetime.strptime(horario_limpo, formato)
                    return dt.strftime('%H:%M')
                except ValueError:
                    continue
            
            return horario_limpo
            
        except Exception:
            return horario or '00:00'
    
    def _horarios_iguais(self, horario1: str, horario2: str) -> bool:
        """
        Compara dois horários ignorando segundos.
        """
        try:
            norm1 = self._normalizar_horario(horario1)
            norm2 = self._normalizar_horario(horario2)
            return norm1 == norm2
        except Exception:
            return str(horario1).strip() == str(horario2).strip()
        
    def extrair_dados_pdf(self, arquivo_pdf) -> Dict:
        """
        Extrai dados de ponto do arquivo PDF.
        """
        try:
            dados_extraidos = {}
            
            with pdfplumber.open(arquivo_pdf) as pdf:
                for pagina in pdf.pages:
                    texto = pagina.extract_text()
                    
                    linhas = texto.split('\n')
                    
                    for linha in linhas:
                        if re.search(r'\d{2}/\d{2}\s+\w+', linha):
                            self._processar_linha_ponto(linha, dados_extraidos)
            
            self.dados_ponto_pdf = dados_extraidos
            return dados_extraidos
            
        except Exception as e:
            st.error(f"Erro ao extrair dados do PDF: {str(e)}")
            return {}
    
    def _processar_linha_ponto(self, linha: str, dados_extraidos: Dict):
        """
        Processa uma linha individual de ponto do PDF.
        """
        try:
            padrao = r'(\d{2}/\d{2})\s+(\w+(?:-\w+)?)\s+(.*)'
            match = re.match(padrao, linha.strip())
            
            if match:
                data_str, dia_semana, resto_linha = match.groups()
                data_completa = f"{data_str}/2025"
                
                # Determinar status baseado no dia da semana e conteúdo
                status = self._determinar_status_dia(dia_semana, resto_linha)
                
                if 'FERIADO' in resto_linha or resto_linha.strip() == '-':
                    dados_extraidos[data_completa] = {
                        'dia_semana': dia_semana,
                        'status': status,
                        'horarios': [],
                        'horas_trabalhadas': '00:00',
                        'saldo': '00:00',
                        'marcacoes_manuais': 0
                    }
                    return
                
                horarios, horas_trabalhadas, marcacoes_manuais = self._extrair_horarios_linha(resto_linha)
                horas_trabalhadas_norm = self._normalizar_horario(horas_trabalhadas)
                _, saldo = self._extrair_horas_previstas_saldo(resto_linha)
                saldo_norm = self._normalizar_horario(saldo)
                
                dados_extraidos[data_completa] = {
                    'dia_semana': dia_semana,
                    'status': status,
                    'horarios': horarios,
                    'horas_trabalhadas': horas_trabalhadas_norm,
                    'saldo': saldo_norm,
                    'marcacoes_manuais': marcacoes_manuais
                }
                
        except Exception as e:
            st.warning(f"⚠️ Erro ao processar linha: {linha[:50]}... - {str(e)}")
    
    def _determinar_status_dia(self, dia_semana: str, resto_linha: str) -> str:
        """
        Determina o status do dia baseado no dia da semana e conteúdo.
        """
        dia_semana_lower = dia_semana.lower()
        
        # Verificar se é feriado
        if 'FERIADO' in resto_linha.upper():
            return 'FERIADO'
        
        # Verificar se é final de semana
        if any(final_semana in dia_semana_lower for final_semana in ['sabado', 'sábado', 'domingo']):
            return 'FINAL DE SEMANA'
        
        # Verificar se tem registros de trabalho
        if resto_linha.strip() == '-' or not resto_linha.strip():
            return 'SEM REGISTRO'
        
        # Se tem horários, é dia trabalhado
        if re.search(r'\d{2}:\d{2}', resto_linha):
            return 'TRABALHADO'
        
        return 'SEM REGISTRO'
    
    def _extrair_horas_previstas_saldo(self, linha: str) -> Tuple[str, str]:
        """
        Extrai horas previstas e saldo da linha do PDF.
        """
        try:
            linha_limpa = re.sub(r'\(m\)', '', linha)
            padrao_completo = r'(?:\d{2}:\d{2}\s*)+\|\s*([\d:]+)\s+([\d:]+)(?:\s+([\d:]+))?'
            match = re.search(padrao_completo, linha_limpa)
            
            if match:
                horas_previstas = match.group(2).strip()
                saldo = match.group(3).strip() if match.group(3) else "00:00"
                return horas_previstas, saldo
            
            return "08:00", "00:00"
            
        except Exception:
            return "08:00", "00:00"
    
    def _extrair_horarios_linha(self, linha: str) -> Tuple[List[str], str, int]:
        """
        Extrai horários, horas trabalhadas e contagem de marcações manuais.
        """
        horarios = []
        marcacoes_manuais = linha.count('(m)')
        linha_limpa = re.sub(r'\(m\)', '', linha)
        
        try:
            # Método 1: Buscar padrão completo com pipes
            padrao_pipes = r'((?:\d{2}:\d{2}[\s|]*)+)\|\s*([\d:]+)\s+([\d:]+)(?:\s+([\d:]+))?'
            match_pipes = re.search(padrao_pipes, linha_limpa)
            
            if match_pipes:
                pontos_str = match_pipes.group(1).strip()
                horas_trabalhadas = match_pipes.group(2).strip()
                
                horarios_encontrados = re.findall(r'\d{2}:\d{2}', pontos_str)
                
                for i in range(0, len(horarios_encontrados), 2):
                    if i + 1 < len(horarios_encontrados):
                        entrada = horarios_encontrados[i]
                        saida = horarios_encontrados[i + 1]
                        horarios.append(f"{entrada}-{saida}")
                
                return horarios, horas_trabalhadas, marcacoes_manuais
            
            # Método 2: Buscar todos os tempos e identificar posições
            todos_horarios = re.findall(r'\d{2}:\d{2}', linha_limpa)
            
            if len(todos_horarios) >= 3:
                if len(todos_horarios) >= 4:
                    horas_trabalhadas = todos_horarios[-3]
                    horarios_pontos = todos_horarios[:-3]
                elif len(todos_horarios) == 3:
                    if todos_horarios[0] < todos_horarios[1]:
                        horarios_pontos = todos_horarios[:-1]
                        horas_trabalhadas = todos_horarios[-1]
                    else:
                        horas_trabalhadas = todos_horarios[-2]
                        horarios_pontos = todos_horarios[:-2]
                else:
                    horas_trabalhadas = todos_horarios[-2]
                    horarios_pontos = todos_horarios[:-2]
                
                for i in range(0, len(horarios_pontos), 2):
                    if i + 1 < len(horarios_pontos):
                        entrada = horarios_pontos[i]
                        saida = horarios_pontos[i + 1]
                        horarios.append(f"{entrada}-{saida}")
                
                return horarios, horas_trabalhadas, marcacoes_manuais
            
            # Método 3: Fallback - padrão específico
            padrao_especifico = r'(\d{2}:\d{2})\s+(\d{2}:\d{2})\s*\|\s*(\d{2}:\d{2})\s+(\d{2}:\d{2})\s*\|\s*(\d{2}:\d{2})\s+(\d{2}:\d{2})'
            match_especifico = re.search(padrao_especifico, linha_limpa)
            
            if match_especifico:
                entrada1 = match_especifico.group(1)
                saida1 = match_especifico.group(2) 
                entrada2 = match_especifico.group(3)
                saida2 = match_especifico.group(4)
                horas_trabalhadas = match_especifico.group(5)
                
                horarios = [f"{entrada1}-{saida1}", f"{entrada2}-{saida2}"]
                
                return horarios, horas_trabalhadas, marcacoes_manuais
            
            return [], "00:00", marcacoes_manuais
            
        except Exception:
            return [], "00:00", marcacoes_manuais
    
    def extrair_dados_excel(self, arquivo_excel) -> Dict:
        """
        Extrai dados de apropriação do arquivo Excel.
        """
        try:
            df = pd.read_excel(arquivo_excel, sheet_name=0)
            
            header_row = None
            for idx, row in df.iterrows():
                if 'Data' in str(row.iloc[0]):
                    header_row = idx
                    break
            
            if header_row is None:
                st.error("Não foi possível encontrar o cabeçalho no arquivo Excel")
                return {}
            
            df = pd.read_excel(arquivo_excel, sheet_name=0, header=header_row)
            
            dados_extraidos = {}
            ultima_data_valida = None
            
            for idx, row in df.iterrows():
                data_cell = str(row.iloc[0]).strip()
                
                if self._validar_data(data_cell):
                    ultima_data_valida = self._normalizar_data(data_cell)
                    
                    if ultima_data_valida not in dados_extraidos:
                        dados_extraidos[ultima_data_valida] = {
                            'dia_semana': str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else '',
                            'atividades': [],
                            'total_dia': ''
                        }
                
                if ultima_data_valida and pd.notna(row.iloc[2]) and str(row.iloc[2]).strip():
                    inicio = str(row.iloc[2]).strip()
                    termino = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                    total_atividade = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
                    total_dia = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
                    atividade = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ''
                    
                    dados_extraidos[ultima_data_valida]['atividades'].append({
                        'inicio': inicio,
                        'termino': termino,
                        'total_atividade': total_atividade,
                        'atividade': atividade
                    })
                    
                    if total_dia and not dados_extraidos[ultima_data_valida]['total_dia']:
                        dados_extraidos[ultima_data_valida]['total_dia'] = self._normalizar_horario(total_dia)
            
            dados_limpos = {
                data: info for data, info in dados_extraidos.items()
                if info['atividades'] and any(a['inicio'] for a in info['atividades'])
            }
            
            self.dados_apropriacao_excel = dados_limpos
            return dados_limpos
            
        except Exception as e:
            st.error(f"Erro ao extrair dados do Excel: {str(e)}")
            return {}
    
    def _validar_data(self, data_str: str) -> bool:
        """Valida se uma string representa uma data válida."""
        try:
            padrao = r'\d{1,2}/\d{1,2}/\d{2,4}'
            return bool(re.match(padrao, data_str.strip()))
        except:
            return False
    
    def _normalizar_data(self, data_str: str) -> str:
        """Normaliza formato de data para DD/MM/YYYY."""
        try:
            for formato in ['%d/%m/%Y', '%d/%m/%y']:
                try:
                    data_obj = datetime.strptime(data_str.strip(), formato)
                    return data_obj.strftime('%d/%m/%Y')
                except:
                    continue
            return data_str.strip()
        except:
            return data_str.strip()
    
    def comparar_relatorios(self) -> Dict:
        """
        Compara os relatórios de ponto e apropriação.
        """
        resultado = {
            'divergencias_quantitativas': [],
            'divergencias_por_dia': [],
            'divergencias_horarios': [],
            'marcacoes_manuais_resumo': {},
            'resumo_geral': {}
        }
        
        try:
            datas_ponto = set(self.dados_ponto_pdf.keys())
            datas_apropriacao = set(self.dados_apropriacao_excel.keys())
            todas_datas = datas_ponto.union(datas_apropriacao)
            
            if len(datas_ponto) != len(datas_apropriacao):
                resultado['divergencias_quantitativas'].append({
                    'tipo': 'Quantidade de dias diferentes',
                    'ponto': len(datas_ponto),
                    'apropriacao': len(datas_apropriacao),
                    'diferenca': abs(len(datas_ponto) - len(datas_apropriacao))
                })
            
            datas_so_ponto = datas_ponto - datas_apropriacao
            if datas_so_ponto:
                resultado['divergencias_quantitativas'].append({
                    'tipo': 'Datas apenas no ponto',
                    'datas': sorted(list(datas_so_ponto))
                })
            
            datas_so_apropriacao = datas_apropriacao - datas_ponto
            if datas_so_apropriacao:
                resultado['divergencias_quantitativas'].append({
                    'tipo': 'Datas apenas na apropriação',
                    'datas': sorted(list(datas_so_apropriacao))
                })
            
            for data in sorted(todas_datas):
                self._comparar_dia_especifico(data, resultado)
            
            total_marcacoes = sum(
                info.get('marcacoes_manuais', 0) 
                for info in self.dados_ponto_pdf.values()
            )
            resultado['marcacoes_manuais_resumo'] = {
                'total': total_marcacoes,
                'por_dia': {
                    data: info.get('marcacoes_manuais', 0)
                    for data, info in self.dados_ponto_pdf.items()
                    if info.get('marcacoes_manuais', 0) > 0
                }
            }
            
            resultado['resumo_geral'] = {
                'total_dias_ponto': len(datas_ponto),
                'total_dias_apropriacao': len(datas_apropriacao),
                'dias_comuns': len(datas_ponto.intersection(datas_apropriacao)),
                'total_divergencias': len(resultado['divergencias_por_dia']) + len(resultado['divergencias_horarios']),
                'total_marcacoes_manuais': total_marcacoes
            }
            
        except Exception as e:
            st.error(f"Erro na comparação: {str(e)}")
        
        return resultado
    
    def _comparar_dia_especifico(self, data: str, resultado: Dict):
        """Compara dados de um dia específico entre os relatórios."""
        try:
            info_ponto = self.dados_ponto_pdf.get(data, {})
            info_apropriacao = self.dados_apropriacao_excel.get(data, {})
            
            if not info_ponto and info_apropriacao:
                resultado['divergencias_por_dia'].append({
                    'data': data,
                    'tipo': 'Dia existe apenas na apropriação',
                    'total_apropriacao': info_apropriacao.get('total_dia', ''),
                    'atividades': len(info_apropriacao.get('atividades', []))
                })
                return
            
            if info_ponto and not info_apropriacao:
                resultado['divergencias_por_dia'].append({
                    'data': data,
                    'tipo': 'Dia existe apenas no ponto',
                    'horas_trabalhadas_ponto': info_ponto.get('horas_trabalhadas', ''),
                    'status': info_ponto.get('status', ''),
                    'marcacoes_manuais': info_ponto.get('marcacoes_manuais', 0)
                })
                return
            
            if info_ponto and info_apropriacao:
                horas_trabalhadas_ponto = info_ponto.get('horas_trabalhadas', '00:00')
                total_apropriacao = info_apropriacao.get('total_dia', '00:00')
                
                if not self._horarios_iguais(horas_trabalhadas_ponto, total_apropriacao):
                    resultado['divergencias_por_dia'].append({
                        'data': data,
                        'tipo': 'Divergência entre horas trabalhadas (PDF) e total apropriado (Excel)',
                        'horas_trabalhadas_ponto': horas_trabalhadas_ponto,
                        'total_apropriacao': total_apropriacao,
                        'saldo_ponto': info_ponto.get('saldo', ''),
                        'marcacoes_manuais': info_ponto.get('marcacoes_manuais', 0)
                    })
                
                self._comparar_horarios_dia(data, info_ponto, info_apropriacao, resultado)
        
        except Exception as e:
            st.warning(f"Erro ao comparar dia {data}: {str(e)}")
    
    def _comparar_horarios_dia(self, data: str, info_ponto: Dict, info_apropriacao: Dict, resultado: Dict):
        """Compara horários de entrada/saída com início/fim das atividades."""
        try:
            horarios_ponto = info_ponto.get('horarios', [])
            atividades = info_apropriacao.get('atividades', [])
            
            if not horarios_ponto or not atividades:
                return
            
            if horarios_ponto:
                primeiro_periodo = horarios_ponto[0].split('-')
                ultimo_periodo = horarios_ponto[-1].split('-')
                
                entrada_ponto = primeiro_periodo[0] if len(primeiro_periodo) > 0 else ''
                saida_ponto = ultimo_periodo[1] if len(ultimo_periodo) > 1 else ''
            
            if atividades:
                primeira_atividade = min(atividades, key=lambda x: x.get('inicio', '23:59'))
                ultima_atividade = max(atividades, key=lambda x: x.get('termino', '00:00'))
                
                entrada_apropriacao = primeira_atividade.get('inicio', '')
                saida_apropriacao = ultima_atividade.get('termino', '')
            
            if entrada_ponto and entrada_apropriacao:
                if not self._horarios_proximos(entrada_ponto, entrada_apropriacao):
                    resultado['divergencias_horarios'].append({
                        'data': data,
                        'tipo': 'Divergência horário entrada',
                        'ponto': self._normalizar_horario(entrada_ponto),
                        'apropriacao': self._normalizar_horario(entrada_apropriacao)
                    })
            
            if saida_ponto and saida_apropriacao:
                if not self._horarios_proximos(saida_ponto, saida_apropriacao):
                    resultado['divergencias_horarios'].append({
                        'data': data,
                        'tipo': 'Divergência horário saída',
                        'ponto': self._normalizar_horario(saida_ponto),
                        'apropriacao': self._normalizar_horario(saida_apropriacao)
                    })
        
        except Exception as e:
            st.warning(f"Erro ao comparar horários do dia {data}: {str(e)}")
    
    def _horarios_proximos(self, hora1: str, hora2: str, tolerancia_minutos: int = 5) -> bool:
        """Verifica se dois horários estão próximos dentro de uma tolerância."""
        try:
            h1_norm = self._normalizar_horario(hora1)
            h2_norm = self._normalizar_horario(hora2)
            
            if h1_norm == h2_norm:
                return True
            
            h1 = datetime.strptime(h1_norm, '%H:%M')
            h2 = datetime.strptime(h2_norm, '%H:%M')
            diferenca = abs((h1 - h2).total_seconds() / 60)
            return diferenca <= tolerancia_minutos
        except:
            return self._horarios_iguais(hora1, hora2)

def gerar_relatorio_detalhado_xlsx(resultado: Dict, dados_ponto: Dict, dados_apropriacao: Dict) -> bytes:
    """
    Gera um relatório detalhado focado na comparação principal em formato Excel.
    """
    if not OPENPYXL_AVAILABLE:
        st.error("❌ openpyxl não está disponível. Instale com: pip install openpyxl")
        return b""
    
    try:
        # Verificações básicas de dados
        if not dados_ponto and not dados_apropriacao:
            st.error("❌ Nenhum dado disponível para gerar o relatório")
            return b""
        
        if not resultado:
            st.warning("⚠️ Dados de resultado não disponíveis, gerando relatório básico")
            resultado = {'resumo_geral': {}, 'divergencias_por_dia': [], 'divergencias_horarios': []}
        
        output = io.BytesIO()
        comparador_temp = PontoComparador()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # 1. ABA PRINCIPAL - RELATÓRIO DETALHADO
            st.info("📊 Gerando relatório detalhado...")
            todas_datas = sorted(set(dados_ponto.keys()).union(set(dados_apropriacao.keys())))
            
            # Verificar se temos datas para processar
            if not todas_datas:
                st.warning("⚠️ Nenhuma data encontrada nos dados")
                todas_datas = [datetime.now().strftime('%d/%m/%Y')]  # Data de fallback
            
            dados_detalhados = []
            
            for data in todas_datas:
                info_ponto = dados_ponto.get(data, {})
                info_apropriacao = dados_apropriacao.get(data, {})
                
                horas_trabalhadas_ponto = info_ponto.get('horas_trabalhadas', '')
                horas_previstas = info_ponto.get('horas_previstas', '')
                saldo_ponto = info_ponto.get('saldo', '')
                status_ponto = info_ponto.get('status', 'SEM_REGISTRO')
                marcacoes_manuais = info_ponto.get('marcacoes_manuais', 0)
                horarios_ponto = ' | '.join(info_ponto.get('horarios', []))
                
                total_apropriacao = info_apropriacao.get('total_dia', '')
                qtd_atividades = len(info_apropriacao.get('atividades', []))
                dia_semana = info_apropriacao.get('dia_semana', '') or info_ponto.get('dia_semana', '')
                
                # Determinar status da comparação
                diferenca_horas = ''
                observacao_divergencia = 'SEM DIVERGÊNCIA'
                
                if horas_trabalhadas_ponto and total_apropriacao:
                    if not comparador_temp._horarios_iguais(horas_trabalhadas_ponto, total_apropriacao):
                        diferenca_horas = f"Diferença: {horas_trabalhadas_ponto} ≠ {total_apropriacao}"
                        observacao_divergencia = 'DIVERGÊNCIA DE HORAS'
                elif horas_trabalhadas_ponto and not total_apropriacao:
                    observacao_divergencia = 'PONTO SEM APROPRIAÇÃO'
                elif not horas_trabalhadas_ponto and total_apropriacao:
                    observacao_divergencia = 'APROPRIAÇÃO SEM PONTO'
                elif not horas_trabalhadas_ponto and not total_apropriacao:
                    observacao_divergencia = 'SEM REGISTRO EM AMBOS'
                
                if marcacoes_manuais > 0:
                    if observacao_divergencia == 'SEM DIVERGÊNCIA':
                        observacao_divergencia = f'OK - {marcacoes_manuais} MARCAÇÃO(ÕES) MANUAL(IS)'
                    else:
                        observacao_divergencia += f' + {marcacoes_manuais} MANUAL(IS)'
                
                # Extrair detalhes das atividades
                detalhes_atividades = []
                if info_apropriacao.get('atividades'):
                    for atividade in info_apropriacao['atividades']:
                        inicio = atividade.get('inicio', '')
                        termino = atividade.get('termino', '')
                        desc = atividade.get('atividade', '')
                        total = atividade.get('total_atividade', '')
                        detalhes_atividades.append(f"{inicio}-{termino}: {desc} ({total})")
                
                atividades_str = ' | '.join(detalhes_atividades) if detalhes_atividades else 'Nenhuma atividade'
                
                dados_detalhados.append({
                    'Data': data,
                    'Dia da Semana': dia_semana,
                    'Status Ponto': status_ponto,
                    'Horas Trabalhadas (PDF)': horas_trabalhadas_ponto,
                    'Horas Previstas': horas_previstas,
                    'Saldo Ponto': saldo_ponto,
                    'Total Apropriação (Excel)': total_apropriacao,
                    'Qtd Atividades': qtd_atividades,
                    'Marcações Manuais': marcacoes_manuais,
                    'Horários (Pontos)': horarios_ponto,
                    'Status Comparação': observacao_divergencia,
                    'Diferença Detalhada': diferenca_horas,
                    'Atividades Detalhadas': atividades_str
                })
            
            # Verificar se conseguimos criar dados
            if not dados_detalhados:
                st.warning("⚠️ Nenhum dado detalhado foi criado, gerando registro de fallback")
                dados_detalhados.append({
                    'Data': datetime.now().strftime('%d/%m/%Y'),
                    'Dia da Semana': 'N/A',
                    'Status Ponto': 'SEM DADOS',
                    'Horas Trabalhadas (PDF)': '00:00',
                    'Saldo Ponto': '00:00',
                    'Total Apropriação (Excel)': '00:00',
                    'Qtd Atividades': 0,
                    'Marcações Manuais': 0,
                    'Horários (Pontos)': 'N/A',
                    'Status Comparação': 'ERRO NA EXTRAÇÃO',
                    'Diferença Detalhada': 'Verifique os arquivos de entrada',
                    'Atividades Detalhadas': 'Nenhuma atividade encontrada',
                    'tipo_formatacao': 'divergencia'
                })
            
            df_detalhado = pd.DataFrame(dados_detalhados)
            
            # Verificar se dados foram criados corretamente
            if df_detalhado.empty:
                st.error("❌ Nenhum dado foi processado para o relatório detalhado")
                return b""
            
            # Remover coluna auxiliar antes de salvar no Excel (se existir)
            df_detalhado_excel = df_detalhado.copy()
            if 'tipo_formatacao' in df_detalhado_excel.columns:
                df_detalhado_excel = df_detalhado_excel.drop('tipo_formatacao', axis=1)
            
            df_detalhado_excel.to_excel(writer, sheet_name='RELATÓRIO DETALHADO', index=False)
            
            # 2. ABA RESUMO EXECUTIVO
            st.info("📊 Gerando resumo executivo...")
            resumo = resultado.get('resumo_geral', {})
            dados_resumo = [
                ['Data/Hora da Análise', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
                ['Critério de Comparação', 'PDF: Coluna TRABALHADAS vs Excel: Total do Dia'],
                ['Sistema de Normalização', 'ATIVO - Segundos ignorados automaticamente'],
                ['Versão do Comparador', '2.0 - Corrigida para compatibilidade'],
                ['', ''],
                ['MÉTRICAS PRINCIPAIS', ''],
                ['Dias no Ponto', resumo.get('total_dias_ponto', 0)],
                ['Dias na Apropriação', resumo.get('total_dias_apropriacao', 0)],
                ['Dias em Comum', resumo.get('dias_comuns', 0)],
                ['Total de Divergências', resumo.get('total_divergencias', 0)],
                ['Total de Marcações Manuais', resumo.get('total_marcacoes_manuais', 0)],
                ['', ''],
                ['RESULTADO DA ANÁLISE', ''],
                ['Status', 'APROVADO' if resumo.get('total_divergencias', 0) == 0 else 'REQUER ATENÇÃO'],
                ['Observações', 'Relatório gerado com normalização de horários e tratamento de erros']
            ]
            
            df_resumo = pd.DataFrame(dados_resumo, columns=['Métrica', 'Valor'])
            
            # Verificar se o DataFrame não está vazio
            if not df_resumo.empty:
                df_resumo.to_excel(writer, sheet_name='Resumo Executivo', index=False)
            else:
                # Criar DataFrame básico se houver problema
                df_resumo_backup = pd.DataFrame([['Erro', 'Não foi possível gerar resumo']], columns=['Métrica', 'Valor'])
                df_resumo_backup.to_excel(writer, sheet_name='Resumo Executivo', index=False)
            
            # 3. ABA DIVERGÊNCIAS ENCONTRADAS
            st.info("📊 Gerando análise de divergências...")
            div_dados = []
            
            # Divergências por dia
            for div in resultado.get('divergencias_por_dia', []):
                div_dados.append({
                    'Tipo': 'Divergência por Dia',
                    'Data': div.get('data', ''),
                    'Descrição': div.get('tipo', ''),
                    'PDF (Trabalhadas)': div.get('horas_trabalhadas_ponto', ''),
                    'Excel (Total)': div.get('total_apropriacao', ''),
                    'Saldo Ponto': div.get('saldo_ponto', ''),
                    'Marcações Manuais': div.get('marcacoes_manuais', 0),
                    'Impacto': 'ALTO' if 'Divergência' in div.get('tipo', '') else 'MÉDIO',
                    'Ação Requerida': 'Verificar registros' if 'Divergência' in div.get('tipo', '') else 'Informativo'
                })
            
            # Divergências de horários
            for div in resultado.get('divergencias_horarios', []):
                div_dados.append({
                    'Tipo': 'Divergência de Horário',
                    'Data': div.get('data', ''),
                    'Descrição': div.get('tipo', ''),
                    'PDF (Trabalhadas)': div.get('ponto', ''),
                    'Excel (Total)': div.get('apropriacao', ''),
                    'Marcações Manuais': 0,
                    'Impacto': 'MÉDIO',
                    'Ação Requerida': 'Verificar tolerância'
                })
            
            if not div_dados:
                div_dados.append({
                    'Tipo': 'Nenhuma Divergência',
                    'Data': 'N/A',
                    'Descrição': '✅ Todos os registros estão consistentes',
                    'PDF (Trabalhadas)': 'OK',
                    'Excel (Total)': 'OK',
                    'Saldo Ponto': 'OK',
                    'Marcações Manuais': 0,
                    'Impacto': 'NENHUM',
                    'Ação Requerida': 'Nenhuma ação necessária'
                })
            
            df_divergencias = pd.DataFrame(div_dados)
            
            # Verificar se o DataFrame não está vazio
            if not df_divergencias.empty:
                df_divergencias.to_excel(writer, sheet_name='Divergências Encontradas', index=False)
            else:
                # Criar DataFrame básico se houver problema
                df_divergencias_backup = pd.DataFrame([{
                    'Tipo': 'Sistema',
                    'Data': 'N/A', 
                    'Descrição': 'Erro ao gerar análise de divergências',
                    'PDF (Trabalhadas)': 'N/A',
                    'Excel (Total)': 'N/A',
                    'Saldo Ponto': 'N/A',
                    'Marcações Manuais': 0,
                    'Impacto': 'N/A',
                    'Ação Requerida': 'Verificar dados de entrada'
                }])
                df_divergencias_backup.to_excel(writer, sheet_name='Divergências Encontradas', index=False)
            
            # APLICAR FORMATAÇÃO PROFISSIONAL
            st.info("🎨 Aplicando formatação profissional...")
            workbook = writer.book
            
            # Verificar se o workbook tem abas
            if not workbook.sheetnames:
                st.error("❌ Erro: Nenhuma aba foi criada no workbook")
                return b""
            
            # Cores e estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar formatação em cada aba
            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    
                    # Verificar se a worksheet tem dados
                    if worksheet.max_row < 1 or worksheet.max_column < 1:
                        st.warning(f"⚠️ Aba '{sheet_name}' parece estar vazia")
                        continue
                    
                    # Formatação do cabeçalho
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # Autoajustar largura das colunas
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if cell.value is not None and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max(max_length + 2, 10), 60)  # Mínimo 10, máximo 60
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formatação específica para RELATÓRIO DETALHADO
                    if sheet_name == 'RELATÓRIO DETALHADO' and dados_detalhados:
                        for row in range(2, len(dados_detalhados) + 2):
                            try:
                                # Pegar o tipo de formatação do registro correspondente
                                idx_registro = row - 2  # Ajustar índice (linha 2 = registro 0)
                                if idx_registro < len(dados_detalhados):
                                    tipo_formatacao = dados_detalhados[idx_registro].get('tipo_formatacao', 'normal')
                                    
                                    # Aplicar formatação baseada no tipo
                                    for col in range(1, 13):  # Colunas A-L (removida coluna Horas Previstas)
                                        cell = worksheet.cell(row=row, column=col)
                                        cell.border = thin_border
                                        
                                        if tipo_formatacao == 'divergencia':
                                            cell.fill = red_fill
                                        elif tipo_formatacao == 'manual':
                                            cell.fill = yellow_fill
                                        # Para 'normal' (SEM DIVERGÊNCIA), não aplica cor de fundo
                                else:
                                    # Fallback: aplicar apenas bordas
                                    for col in range(1, 13):
                                        cell = worksheet.cell(row=row, column=col)
                                        cell.border = thin_border
                            except Exception as e:
                                # Se houver erro, aplicar apenas bordas
                                for col in range(1, 13):
                                    try:
                                        cell = worksheet.cell(row=row, column=col)
                                        cell.border = thin_border
                                    except:
                                        pass
                    
                    # Aplicar bordas nas outras abas
                    else:
                        try:
                            for row in worksheet.iter_rows():
                                for cell in row:
                                    if cell.value is not None:
                                        cell.border = thin_border
                        except Exception:
                            pass
                
                except Exception as e:
                    st.warning(f"⚠️ Erro ao formatar aba '{sheet_name}': {str(e)}")
                    continue
        
        # Verificar se o workbook foi criado corretamente
        output.seek(0)
        excel_data = output.getvalue()
        
        if len(excel_data) == 0:
            st.error("❌ Arquivo Excel gerado está vazio")
            return b""
        
        st.success("✅ Relatório Excel gerado com sucesso!")
        return excel_data
        
    except Exception as e:
        st.error(f"Erro ao gerar relatório Excel: {str(e)}")
        import traceback
        error_details = traceback.format_exc()
        st.error(f"Detalhes do erro: {error_details}")
        
        # Tentar gerar um relatório mínimo de emergência
        try:
            st.info("🔄 Tentando gerar relatório básico de emergência...")
            output_emergency = io.BytesIO()
            
            with pd.ExcelWriter(output_emergency, engine='openpyxl') as writer_emergency:
                # Criar pelo menos uma aba básica
                dados_emergencia = [
                    ['Data', 'Status', 'Observação'],
                    [datetime.now().strftime('%d/%m/%Y'), 'ERRO', 'Erro na geração do relatório principal'],
                    ['', '', 'Verifique os dados de entrada e tente novamente'],
                    ['', '', f'Erro técnico: {str(e)[:100]}...']
                ]
                
                df_emergency = pd.DataFrame(dados_emergencia[1:], columns=dados_emergencia[0])
                df_emergency.to_excel(writer_emergency, sheet_name='Relatório de Erro', index=False)
            
            output_emergency.seek(0)
            return output_emergency.getvalue()
            
        except Exception as e2:
            st.error(f"❌ Falha crítica: não foi possível gerar nem o relatório de emergência: {str(e2)}")
            return b""

def main():
    
    # Upload de arquivos
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📄 Folha de Ponto (PDF)")
        arquivo_pdf = st.file_uploader(
            "Selecione o arquivo PDF",
            type=['pdf'],
            help="Arquivo PDF contendo os registros de ponto"
        )
    
    with col2:
        st.subheader("📊 Apropriação (Excel)")
        arquivo_excel = st.file_uploader(
            "Selecione o arquivo Excel",
            type=['xlsx', 'xls'],
            help="Arquivo Excel com o relatório de apropriação"
        )
    
    # Processar arquivos se ambos foram carregados
    if arquivo_pdf and arquivo_excel:
        
        # Criar instância do comparador
        comparador = PontoComparador()
        
        # Processar arquivos automaticamente
        with st.spinner("🔄 Processando arquivos..."):
            st.info("📄 Extraindo dados da folha de ponto...")
            dados_ponto = comparador.extrair_dados_pdf(arquivo_pdf)
            
            st.info("📊 Extraindo dados da apropriação...")
            dados_apropriacao = comparador.extrair_dados_excel(arquivo_excel)
            
            if dados_ponto and dados_apropriacao:
                st.success("✅ Arquivos processados com sucesso!")
                
                # Realizar comparação
                st.info("🔍 Comparando relatórios...")
                resultado = comparador.comparar_relatorios()
                
                # Exibir resultados
                exibir_resultados(resultado, dados_ponto, dados_apropriacao)
            else:
                st.error("❌ Erro ao processar um ou ambos os arquivos.")
                
                if not dados_ponto:
                    st.error("❌ Falha na extração do PDF. Verifique se o formato está correto.")
                if not dados_apropriacao:
                    st.error("❌ Falha na extração do Excel. Verifique se o formato está correto.")
    else:
        # Mostrar instruções quando não há arquivos
        st.info("👆 **Faça upload dos dois arquivos acima para começar a análise**")
        
        st.markdown("### 🎯 Exemplo do Resultado")
        st.markdown("Após o upload, você verá:")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Dias no Ponto", "23", help="Total de dias encontrados no PDF")
        with col2:
            st.metric("Dias na Apropriação", "23", help="Total de dias no Excel")
        with col3:
            st.metric("Divergências", "2", delta="-2", help="Divergências encontradas")
        with col4:
            st.metric("Marcações Manuais", "5", help="Marcações manuais identificadas")
        
        st.markdown("**E um botão para baixar o relatório Excel completo com análise detalhada!**")

def exibir_resultados(resultado: Dict, dados_ponto: Dict, dados_apropriacao: Dict):
    """Exibe os resultados da comparação de forma simplificada."""
    
    st.markdown("---")
    st.header("📋 Resultados da Análise")
    
    # Resumo geral em métricas
    resumo = resultado.get('resumo_geral', {})
    
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Dias no Ponto", resumo.get('total_dias_ponto', 0))
    with col2:
        st.metric("Dias na Apropriação", resumo.get('total_dias_apropriacao', 0))
    with col3:
        st.metric("Divergências", resumo.get('total_divergencias', 0), 
                 delta=f"-{resumo.get('total_divergencias', 0)}" if resumo.get('total_divergencias', 0) > 0 else "0")
    with col4:
        st.metric("Marcações Manuais", resumo.get('total_marcacoes_manuais', 0))
    with col5:
         # BOTÃO QUE JÁ GERA E BAIXA AUTOMATICAMENTE
        if st.button("📊 **BAIXAR RELATÓRIO EXCEL**", type="primary", use_container_width=True):
            with st.spinner("🔄 Gerando relatório Excel..."):
                relatorio_excel = gerar_relatorio_detalhado_xlsx(resultado, dados_ponto, dados_apropriacao)
                
                if relatorio_excel and len(relatorio_excel) > 0:
                    # Download automático
                    st.download_button(
                        label="⬇️ Clique aqui para baixar",
                        data=relatorio_excel,
                        file_name=f"relatorio_ponto_detalhado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.success("✅ Relatório gerado com sucesso! Use o botão acima para baixar.")
                else:
                    st.error("❌ Erro ao gerar o relatório Excel. Tente novamente.")
    
    # Status geral
    if resumo.get('total_divergencias', 0) == 0:
        st.success("🎉 **Parabéns!** Não foram encontradas divergências entre os relatórios.")
        status_cor = "success"
    else:
        st.warning(f"⚠️ Foram encontradas **{resumo.get('total_divergencias', 0)}** divergências que precisam de atenção.")
        status_cor = "warning"
    
   
    
    # Prévia dos resultados
    st.markdown("### 📊 Resultados")
    
    # Criar prévia simplificada
    todas_datas = sorted(set(dados_ponto.keys()).union(set(dados_apropriacao.keys())))
    dados_previa = []
    
    comparador_temp = PontoComparador()
    
    # Mostrar até 15 registros na prévia
    for data in todas_datas[:15]:
        info_ponto = dados_ponto.get(data, {})
        info_apropriacao = dados_apropriacao.get(data, {})
        
        horas_trabalhadas = info_ponto.get('horas_trabalhadas', '00:00')
        total_apropriacao = info_apropriacao.get('total_dia', '00:00')
        status_ponto = info_ponto.get('status', 'SEM REGISTRO')
        marcacoes_manuais = info_ponto.get('marcacoes_manuais', 0)
        
        # Determinar status visual
        if horas_trabalhadas and total_apropriacao:
            if not comparador_temp._horarios_iguais(horas_trabalhadas, total_apropriacao):
                status_visual = "🔴 DIVERGÊNCIA"
                observacao = f'{horas_trabalhadas} ≠ {total_apropriacao}'
            else:
                if marcacoes_manuais > 0:
                    status_visual = f"🟡 MANUAL ({marcacoes_manuais})"
                    observacao = 'Registros conferem'
                else:
                    status_visual = "🟢 OK"
                    observacao = 'Perfeito'
        elif horas_trabalhadas and not total_apropriacao:
            status_visual = "🔴 SÓ NO PONTO"
            observacao = 'Falta apropriação'
        elif not horas_trabalhadas and total_apropriacao:
            status_visual = "🔴 SÓ NA APROPRIAÇÃO"
            observacao = 'Falta ponto'
        else:
            status_visual = "⚪ SEM DADOS"
            observacao = 'Ambos vazios'
        
        dados_previa.append({
            'Data': data,
            'Status Ponto': status_ponto,
            'PDF (Trabalhadas)': horas_trabalhadas or '00:00',
            'Excel (Total)': total_apropriacao or '00:00',
            'Status': status_visual,
            'Observação': observacao
        })
    
    if dados_previa:
        df_previa = pd.DataFrame(dados_previa)
        st.dataframe(df_previa, use_container_width=True, height=400)
    else:
        st.warning("Nenhum dado encontrado para prévia.")
    
    
        
    
    
    
   
 

if __name__ == "__main__":
    main()